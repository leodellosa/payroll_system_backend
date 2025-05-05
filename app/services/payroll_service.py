from typing import List, Optional
from app.models.payroll import Payroll
from app.schemas.payroll import PayrollSchema, PayrollResponse,PayrollCreate,PayrollUpdate
from sqlalchemy.orm import Session
from sqlalchemy.inspection import inspect
from app.models.employee import Employee
from sqlalchemy.exc import IntegrityError
from fastapi import HTTPException
from weasyprint import HTML, CSS
from jinja2 import Environment, FileSystemLoader
import os
from decimal import Decimal
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font
from fastapi.responses import StreamingResponse

def get_employee_or_404(db: Session, employee_id: int):
    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail=f"Employee with ID {employee_id} not found")
    return employee

def get_all_payrolls(db: Session) -> List[PayrollResponse]:
    payrolls = db.query(Payroll).all()
    return [
        PayrollResponse(**{c.key: getattr(payroll, c.key) for c in inspect(Payroll).mapper.column_attrs})
        for payroll in payrolls
    ]

def generate_payroll(db: Session, payroll_data: PayrollCreate, employee_id: int) -> dict:
    get_employee_or_404(db, employee_id)
    try:
        new_payroll = Payroll(**payroll_data.model_dump())
        db.add(new_payroll)
        db.commit()
        db.refresh(new_payroll)

        return {"success": True, "payroll": new_payroll}
    except IntegrityError as e:
        db.rollback()
        return {"success": False, "error": "Payroll record already exists for this employee on that date.", "code": 400}
    except Exception as e:
        db.rollback()
        return {"success": False, "error": str(e), "code": 500}
    
def update_payroll(db: Session, payroll_id: int, payroll_data: PayrollUpdate) -> dict:
    payroll = db.query(Payroll).filter(Payroll.id == payroll_id).first()
    if not payroll:
        return {"success": False, "error": "Payroll record not found", "code": 404}
    try:
        update_data = payroll_data.model_dump()
        for key, value in update_data.items():
            setattr(payroll, key, value)

        db.commit()
        db.refresh(payroll)
        return {"success": True, "payroll": payroll}
    except IntegrityError:
        db.rollback()
        return {"success": False, "error": "Payroll record already exists for this employee on that date.", "code": 400}
    except Exception as e:
        db.rollback()
        return {"success": False, "error": f"An error occurred: {str(e)}", "code": 500}

def delete_payroll(db: Session, payroll_id: int) -> dict:
    payroll = db.query(Payroll).filter(Payroll.id == payroll_id).first()
    if not payroll:
        return {"success": False, "error": "Payroll record not found", "code": 404}
    try:
        db.delete(payroll)
        db.commit()
        return {"success": True, "message": "Payroll record deleted successfully."}
    except Exception as e:
        db.rollback()
        return {"success": False, "error": f"An error occurred: {str(e)}", "code": 500}   

def get_payroll_summary(
    db: Session,
    employee_id: int,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> dict:
   
    get_employee_or_404(db, employee_id)
    try:
        payrolls_query = db.query(Payroll).filter(Payroll.employee_id == employee_id)
        if start_date:
            payrolls_query = payrolls_query.filter(Payroll.date >= start_date)
        if end_date:
            payrolls_query = payrolls_query.filter(Payroll.date <= end_date)
        payrolls = payrolls_query.all()

        if not payrolls:
            return {"success": False, "error": "No payroll records found.", "code": 404}   

        payrolls_data = [
            PayrollResponse(**{c.key: getattr(payroll, c.key) for c in inspect(Payroll).mapper.column_attrs})
            for payroll in payrolls
        ]
        return {"success": True, "payrolls": payrolls_data}
    except Exception as e:
        db.rollback()
        return {"success": False, "error": str(e), "code": 500}
    

def generate_payslip_pdf(employee_id: int, db: Session) -> dict:

    employee = get_employee_or_404(db, employee_id)
    try:
        payrolls = db.query(Payroll).filter(Payroll.employee_id == employee_id).all()
        if not payrolls:
            return {"success": False, "error": "No payroll records found for this employee.", "code": 404}

        # Payroll summary
        total_hours_worked = sum(p.total_hours_worked or Decimal(0) for p in payrolls)
        total_overtime_pay = sum(p.overtime_pay or Decimal(0) for p in payrolls)
        total_night_diff = sum(p.night_differential_pay or Decimal(0) for p in payrolls)
        total_deductions = sum(p.deductions or Decimal(0) for p in payrolls)
        allowance = sum(p.allowance or Decimal(0) for p in payrolls)
        gross_salary = sum(p.subtotal or Decimal(0) for p in payrolls)
        net_salary = sum(p.net_salary or Decimal(0) for p in payrolls)
        pay_period_from = min(p.date for p in payrolls)
        pay_period_to = max(p.date for p in payrolls)
        daily_rate = employee.salary
        current_date = datetime.now()

        # Render HTML with Jinja2
        print("Current working directory:", os.getcwd())
        template_env = Environment(loader=FileSystemLoader("app/templates"))
        template = template_env.get_template("payroll_payslip.html") 
        html_content = template.render(
            employee=employee,
            payrolls=payrolls,
            total_hours_worked=total_hours_worked,
            total_overtime_pay=total_overtime_pay,
            total_night_diff=total_night_diff,
            total_deductions=total_deductions,
            allowance=allowance,
            gross_salary=gross_salary,
            net_salary=net_salary,
            pay_period_from=pay_period_from,
            pay_period_to=pay_period_to,
            current_date=current_date,
            daily_rate=daily_rate
        )
        
        print(employee)
        # Generate and save PDF file
        output_path = f"pdf/payslip_{employee.first_name}_{employee.last_name}_{current_date.strftime('%Y%m%d_%H%M%S')}.pdf"
        os.makedirs("pdf", exist_ok=True)
        base_url = os.path.abspath("app/static")
        landscape_css = CSS(string='''
        @page {
            size: A4 landscape;
            margin: 1cm;
        }
        ''')
        pdf_bytes = HTML(string=html_content, base_url=base_url).write_pdf(stylesheets=[landscape_css])
        with open(output_path, "wb") as f:
            f.write(pdf_bytes)

        return {
            "success": True,
            "message": "Payslip PDF generated successfully.",
            "filename": os.path.basename(output_path),
            "path": output_path
        }

    except Exception as e:
        return {
            "success": False,
            "error": f"An error occurred during PDF generation: {str(e)}"
        }
    

def generate_payslip_excel(db: Session, employee_id: int) -> dict:
    employee = get_employee_or_404(db, employee_id)
    payrolls = db.query(Payroll).filter(Payroll.employee_id == employee_id).all()

    if not payrolls:
        raise HTTPException(status_code=404, detail="No payroll records found for this employee.")
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Payslip"

        logo_path = os.path.join("app", "static", "images", "company_logo.png")
        try:
            img = ExcelImage(logo_path)
            img.width = 120
            img.height = 80
            ws.add_image(img, 'A1')
        except Exception:
            pass  # Logo optional, continue without crashing

        ws.merge_cells('C1:G1')
        ws['C1'] = "HODREAL FIT-OUT AND CONSTRUCTION"
        ws['C1'].alignment = Alignment(horizontal='left')
        ws['C1'].font = Font(bold=True, size=14)

        ws.merge_cells('C2:J2')
        ws['C2'] = "Batangas City | Contact: 09217292222 | Email: hodrealconstruction@yahoo.com"
        ws['C2'].alignment = Alignment(horizontal='left')
        ws['C2'].font = Font(size=10)

        row_offset = 5
        ws.merge_cells(f'A{row_offset}:F{row_offset}')
        ws[f'A{row_offset}'] = f"Employee: {employee.first_name} {employee.last_name}"
        ws[f'A{row_offset}'].font = Font(bold=True)

        ws.merge_cells(f'A{row_offset + 1}:F{row_offset + 1}')
        ws[f'A{row_offset + 1}'] = f"Position: {employee.position}"
        ws[f'A{row_offset + 1}'].font = Font(bold=True)

        pay_period_from = min([p.date for p in payrolls])
        pay_period_to = max([p.date for p in payrolls])

        ws.merge_cells(f'A{row_offset + 2}:F{row_offset + 2}')
        ws[f'A{row_offset + 2}'] = f"Pay Period: {pay_period_from} - {pay_period_to}"
        ws[f'A{row_offset + 2}'].font = Font(bold=True)

        ws.merge_cells(f'A{row_offset + 3}:F{row_offset + 3}')
        ws[f'A{row_offset + 3}'] = f"Date Generated: {datetime.now().strftime('%Y-%m-%d')}"
        ws[f'A{row_offset + 3}'].font = Font(bold=True)

        daily_rate = employee.salary
        ws.merge_cells(f'A{row_offset + 4}:F{row_offset + 4}')
        ws[f'A{row_offset + 4}'] = f"Daily Rate: {daily_rate}"
        ws[f'A{row_offset + 4}'].font = Font(bold=True)

        row_offset += 7

        headers = ["Date", "Overtime", "Night Diff", "Allowance", "Deductions", "Net Salary"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=row_offset, column=col_num, value=header).font = Font(bold=True)

        for payroll in payrolls:
            row_offset += 1
            ws.append([
                payroll.date.strftime("%Y-%m-%d"),
                round(payroll.overtime_pay, 2),
                round(payroll.night_differential_pay, 2),
                round(payroll.allowance, 2),
                round(payroll.deductions, 2),
                round(payroll.net_salary, 2),
            ])

        total_hours = sum(p.total_hours_worked for p in payrolls)
        total_ot = sum(p.overtime_pay for p in payrolls)
        total_nd = sum(p.night_differential_pay for p in payrolls)
        total_allowance = sum(p.allowance for p in payrolls)
        total_deductions = sum(p.deductions for p in payrolls)
        total_gross = sum(p.subtotal for p in payrolls)
        total_net = sum(p.net_salary for p in payrolls)

        row_offset += 2
        summary_data = [
            ("Total Hours Worked", total_hours),
            ("Total Overtime Pay", total_ot),
            ("Total Night Differential Pay", total_nd),
            ("Total Allowance", total_allowance),
            ("Total Deductions", total_deductions),
            ("Total Gross Salary", total_gross),
            ("Total Net Salary", total_net),
        ]

        for label, value in summary_data:
            ws[f'A{row_offset}'] = f"{label}: {value}"
            row_offset += 1

        folder_path = "excel"
        os.makedirs(folder_path, exist_ok=True)

        # Generate the file name
        filename = f"payslip_{employee.first_name}_{employee.last_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = os.path.join(folder_path, filename)

        # Save the file to the folder
        wb.save(file_path)

        return {
            "success": True,
            "message": "Payslip Excel file generated successfully.",
            "filename": filename,
            "path": file_path
        }

        # Write to buffer
        # output = BytesIO()
        # wb.save(output)
        # output.seek(0)

        # filename = f"payslip_{employee.first_name}_{employee.last_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        # headers = {
        #     "Content-Disposition": f"attachment; filename={filename}"
        # }

        # return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
    except Exception as e:      
        raise HTTPException(status_code=500, detail=f"An error occurred while generating the payslip: {str(e)}")


from fastapi import APIRouter, UploadFile, File, HTTPException, Depends

from io import BytesIO
from typing import List

import pandas as pd



REQUIRED_COLUMNS = [
    'employee_id', 'allowance', 'total_hours_worked', 'overtime_pay',
    'overtime_hour', 'night_differential_pay', 'night_differential_hour',
    'deductions', 'deduction_remarks', 'subtotal', 'net_salary', 'date',
    'time_in', 'time_out', 'project'
]

def fill_missing_values(df: pd.DataFrame) -> pd.DataFrame:
    for column in df.columns:
        for index, value in df[column].items():
            if pd.isna(value):
                if column in ['allowance', 'total_hours_worked',
                              'overtime_pay', 'overtime_hour', 'night_differential_pay',
                              'night_differential_hour', 'deductions', 'subtotal', 'net_salary']:
                    df.at[index, column] = Decimal('0.00')
                else:
                    df.at[index, column] = ''
    return df

async def batch_upload_payroll(db: Session, excel_file: UploadFile = File(...))-> dict:
    """
    Upload an Excel file and batch insert payroll records.
    """
    try:
        content = await excel_file.read()
        df = pd.read_excel(BytesIO(content))

        # Validate required columns
        for column in REQUIRED_COLUMNS:
            if column not in df.columns:
                raise HTTPException(status_code=400, detail=f"Missing required column: {column}")

        df = fill_missing_values(df)

        payroll_records = []

        for _, row in df.iterrows():
            employee = db.query(Employee).filter_by(id=row['employee_id']).first()
            if not employee:
                raise HTTPException(status_code=404, detail=f"Employee with ID {row['employee_id']} does not exist.")

            try:
                time_in = pd.to_datetime(row['time_in'], format='%H:%M:%S').time()
                time_out = pd.to_datetime(row['time_out'], format='%H:%M:%S').time()
                date_value = pd.to_datetime(row['date']).date()

                payroll = Payroll(
                    employee_id=row['employee_id'],
                    allowance=row['allowance'],
                    total_hours_worked=row['total_hours_worked'],
                    overtime_pay=row['overtime_pay'],
                    overtime_hour=row['overtime_hour'],
                    night_differential_pay=row['night_differential_pay'],
                    night_differential_hour=row['night_differential_hour'],
                    deductions=row['deductions'],
                    deduction_remarks=row.get('deduction_remarks', ''),
                    subtotal=row['subtotal'],
                    net_salary=row['net_salary'],
                    date=date_value,
                    time_in=datetime.combine(date_value, time_in),
                    time_out=datetime.combine(date_value, time_out),
                    project=row.get('project', ''),
                    created_at= datetime.now()
                )

                payroll_records.append(payroll)

            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Error processing row: {str(e)}")

        db.bulk_save_objects(payroll_records)
        db.commit()

        return {"success": True, "message": f"{len(payroll_records)} payroll records uploaded."}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")
    
    
def download_payroll_template():
    """
    Generate and download an Excel template for payroll data entry.
    """
    columns = [
        'employee_id', 'allowance', 'total_hours_worked', 'overtime_pay', 
        'overtime_hour', 'night_differential_pay', 'night_differential_hour', 
        'deductions', 'deduction_remarks', 'subtotal', 'net_salary', 'date', 
        'time_in', 'time_out', 'project'
    ]
    data = {col: [''] if col in ['employee_id', 'deduction_remarks', 'project', 'date', 'time_in', 'time_out'] else [0] for col in columns}
    df = pd.DataFrame(data)

    # Write Excel to in-memory buffer
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Payroll Template')

    output.seek(0)

    headers = {
        'Content-Disposition': 'attachment; filename=payroll_template.xlsx'
    }

    return StreamingResponse(output, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)




    